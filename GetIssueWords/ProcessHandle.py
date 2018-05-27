import openpyxl, pandas, os, OperateSSTXml
from functools import reduce
from openpyxl.utils.dataframe import dataframe_to_rows

def loadExcelToDataFrame(excelPath, fillinDataFrame):
    print('read excel ....')
    worksheetNames = ['Issue Word Details', 'Unsure Word Details']
    workBook = openpyxl.load_workbook(excelPath)
    return fillinDataFrame(workBook, worksheetNames)

def loadScriptAndIssueWordToDataFrame(workBook, worksheetNames):
    scriptColumn = []
    issueWordColumn = []

    for sName in worksheetNames:
        workSheet = workBook[sName]
        scriptColumn += [x.value for x in workSheet['D']][1:]
        issueWordColumn += [x.value for x in workSheet['E']][1:]

    df = pandas.DataFrame({'Script':scriptColumn, 'Issue word':issueWordColumn})
    return df


def loadFullUHRSReport(workBook, worksheetNames):
    dfList = []
    for sName in worksheetNames:
        ws = workBook[sName]
        data = ws.values
        cols = next(data)
        df = pandas.DataFrame(data, columns = cols)
        delColumns = ['Domain', 'Voice', 'Audio', 'LE', 'Issue count']
        for delCol in delColumns:
            del df[delCol]

        audioColName = ws['C'][0].value
        audioColvalues = [x.hyperlink.target for x in ws['C'][1:]]
        df[audioColName] = pandas.Series(audioColvalues)
        df['Issue Type'] = pandas.Series([sName]* df.size)
        dfList.append(df)

    return pandas.concat(dfList) 
        
def isEmoji(content):  
    if len(content) != 1:  
        return False  
    if u"\U0001F600" <= content and content <= u"\U0001F64F":  
        return True  
    elif u"\U0001F300" <= content and content <= u"\U0001F5FF":  
        return True  
    elif u"\U0001F680" <= content and content <= u"\U0001F6FF":  
        return True  
    elif u"\U0001F1E0" <= content and content <= u"\U0001F1FF":  
        return True  
    else:  
        return False 

def fliterTheDataFrame(DataFrame):
    print('splite the issue word and script')
    #remove emoji
    DataFrame['Issue word'] = DataFrame['Issue word'].apply(lambda x: '' if isEmoji(x) else x)
    dfRemoveEmpty = DataFrame[DataFrame['Issue word']!='']
    finalDf = dfRemoveEmpty.drop_duplicates(['Script', 'Issue word'])
    return finalDf

def groupbyIssueWord(df):
    def merge(list):    
        return reduce(lambda x, y: x + ', ' + y, list)
    grouped = df.groupby('Issue word')
    grouped2Column = grouped['Script'].agg(merge)
    return grouped2Column
    
def outputWordList(groupRes, outputPath):
    with open(outputPath, 'w', encoding='utf-16') as writer:
        for issueWord in groupRes.index:
            script = groupRes[issueWord]
            writer.write('{0}\t1\t1\t<Examples>: {1}\n'.format(issueWord, script))

def outputToExcel(df, outputExcelPath):
    df['Correct Pronunciation'] = pandas.Series([])
    df['High Skill LE Comments'] = pandas.Series([])
    df=df[['Audio', 'Script', 'Issue word', 'Actual Pronunciation', 'Correct Pronunciation', 'High Skill LE Comments', 'Pronunciation Source',
          'Comment', 'SsmlScript', 'Issue Type']]
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df, header=True, index = False):
        ws.append(r)
    wb.save(outputExcelPath)
        

def StartProcess(model, excelPath):
    if model == 'Excel':
        df = loadExcelToDataFrame(excelPath, loadFullUHRSReport)
        df = fliterTheDataFrame(df)

        print('Get pronuciation from SST')
        pronAndPronsource = OperateSSTXml.GetInfoFromEngine(df['Script'], 'en-US', df['Issue word'])
        pron, pronSource = zip(* pronAndPronsource)
        df['Actual Pronunciation'] = pandas.Series(pron, index= df.index)
        df['Pronunciation Source'] = pandas.Series(pronSource, index= df.index)

        print('Generate final excel')
        outputExcelPath = os.path.splitext(excelPath)[0] + '_to_high_skill.xlsx'
        outputToExcel(df, outputExcelPath)

    elif model == 'Wordlist':
        df = loadExcelToDataFrame(excelPath, loadScriptAndIssueWordToDataFrame)
        df = fliterTheDataFrame(df)
        gb = groupbyIssueWord(df)
        wordlistPath = os.path.splitext(excelPath)[0] + '.txt'

        print('write word list file')
        outputWordList(gb, wordlistPath)